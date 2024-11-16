def anukarah(a):
    import shubhlipi as sh
    from openpyxl import load_workbook

    sht = load_workbook("converts.xlsx").active

    from dattAMsh import mAtrA

    m1 = {}
    for x in mAtrA:
        m1[x] = mAtrA[x]
    mAtrA = m1
    mAtrA["Assamese"] = sh.deepcopy(mAtrA["Bengali"])
    lang_list = a.keys()
    p = {}
    bhAShAmAtrA = {}
    for x in lang_list:
        p[x] = {}
        bhAShAmAtrA[x] = {}

    def swap(js):
        ret = {}
        for key in js:
            ret[js[key]] = key
        return ret

    for lang in mAtrA:
        bhAShAmAtrA[lang] = swap(mAtrA[lang])
    t1 = sh.deepcopy(bhAShAmAtrA["Sanskrit"])
    bhAShAmAtrA["Hindi"] = sh.deepcopy(t1)
    bhAShAmAtrA["Marathi"] = sh.deepcopy(t1)
    bhAShAmAtrA["Konkani"] = sh.deepcopy(t1)
    bhAShAmAtrA["Nepali"] = sh.deepcopy(t1)
    alt_numb = {}
    for lang in a:
        alt_numb[lang] = {}
        for char in a[lang]:
            if len(char) > 1 or char in ["x"]:
                continue
            for varna in a[lang][char]:
                v = a[lang][char][varna]
                if v[0] == varna and lang != "Romanized":
                    continue
                if varna == "AU":
                    continue
                if v[0] in p[lang]:
                    pUrva = p[lang][v[0]][0]
                    if len(pUrva) > len(varna):
                        p[lang][v[0]] = [varna, 2 if v[-1] == 0 else v[-1]]
                    else:
                        continue
                else:
                    if v[0] in bhAShAmAtrA[lang]:
                        p[lang][v[0]] = [bhAShAmAtrA[lang][v[0]], 0]
                    else:
                        p[lang][v[0]] = [varna, 2 if v[-1] == 0 else v[-1]]
    for lang in p:
        code = 2
        p1 = p[lang]
        ak = a[lang]
        if lang == "Romanized":
            p1[ak["."][".A"][0]] = ["A", 2]
            p1[ak["."][".AI"][0]] = ["ai", 2]
            p1[ak["."][".AU"][0]] = ["au", 2]
            p1[ak["a"]["aiI"][0]] = ["ai", 2]
            p1[ak["a"]["auU"][0]] = ["au", 2]
            p1[ak["."][".O"][0]] = ["O", 2]
            p1[ak["."][".E"][0]] = ["E", 2]
            p1[ak["."][".I"][0]] = ["I", 2]
            p1[ak["."][".U"][0]] = ["U", 2]
            p1[ak["."][".R"][0]] = ["R", 2]
            p1[ak["."][".Rz"][0]] = ["Rz", 2]
            p1[ak["."][".RR"][0]] = ["RR", 2]
            p1[ak["."][".G"][0]] = ["G", 2]
            p1[ak["."][".J"][0]] = ["J", 2]
            p1[ak["."][".Y"][0]] = ["Y", 2]
            p1[ak["."][".T"][0]] = ["T", 2]
            p1[ak["."][".D"][0]] = ["D", 2]
            p1[ak["."][".Dz"][0]] = ["Dz", 2]
            p1[ak["."][".Dhz"][0]] = ["Dhz", 2]
            p1[ak["."][".N"][0]] = ["N", 2]
            p1[ak["."][".LR"][0]] = ["LR", 2]
            p1[ak["."][".LRR"][0]] = ["LRR", 2]
            p1[ak["."][".Lz"][0]] = ["Lz", 2]
            p1[ak["."][".L"][0]] = ["L", 2]
            p1[ak["."][".M"][0]] = ["M", 2]
            p1[ak["."][".MM"][0]] = ["MM", 2]
            p1[ak["."][".M1"][0]] = ["M", 2]
            p1[ak["."][".H"][0]] = ["H", 2]
            p1[ak["L"]["Lz"][0]] = ["Lz", 2]
            p1[ak["r"]["rz"][0]] = ["rz", 2]
            p1[ak["n"]["nz"][0]] = ["nz", 2]
            p1[ak["y"]["yz"][0]] = ["yz", 2]
            p1[ak["M"]["M2"][0]] = ["M", 2]
            p1[ak["."][".M1"][0]] = ["M", 2]
            p1["r"] = ["r", 2]
            p1["l"] = ["l", 2]
            p1["m"] = ["m", 2]
            p1["M"] = ["m", 2]
            del p1["z"]
            del p1["\u200c"]
            del p1["\u200d"]
            for x in ak:
                if not (x.isascii() and x.isupper()) or len(x) > 1 or x in "SM":
                    continue
                p1[ak[x][x + "1"][0]] = [x.lower(), 2]
            p1["ṁ"] = ["M", 2]
            continue
        if lang != "Urdu":  # for non-urdu
            code = 1
            p1[ak["Q"]["QQ"][0]] = ["''", 2]
            p1[ak["Y"]["Y"][0]] = ["Y", code]
        p1[ak["f"]["f"][0]] = ["ph", code if lang != "Tamil-Extended" else 3]
        p1[ak["c"]["ch"][0]] = ["ch", code]
        p1[ak["c"]["chh"][0]] = ["chh", code]
        p1[ak["w"]["w"][0]] = ["v", code]
        p1[ak["q"]["qq"][0]] = [".x", 2]
        p1[ak["S"]["SWS"][0]] = 0
        del p1[ak["S"]["SWS"][0]]  # removing swastika
        if "\u200c" in p1:
            del p1["\u200c"]  # removing zreo width joiner
        if (".z" in ak["."] or lang == "Urdu") and lang not in ("Granth", "Siddham"):
            p1[ak["D"]["Dz"][0]] = ["Dz", code]
            p1[ak["D"]["Dhz"][0]] = ["Dhz", code]
            p1[ak["j"]["jz"][0]] = ["jz", code]
        data = {
            "Malayalam": [
                [],
                {
                    "au`0": ["au", 2],
                    "au1`1": ["au", 0],
                },
            ],
            "Brahmi": [
                [],
                {
                    ".A1`0": ["A", 2],
                },
            ],
            "Sinhala": [
                [],
                {
                    "aiI`0": ["au", 2],
                    "auU`0": ["ai", 2],
                    "aiI`1": ["au", 0],
                    "auU`1": ["ai", 0],
                    "G1`0": ["G", 1],
                    "J1`0": ["J", 1],
                    "j1`0": ["jY", 1],
                    "nz`0": ["nd", 1],
                    "mz`0": ["mb", 1],
                    "Nz`0": ["ND", 1],
                },
            ],
            "Sanskrit": [
                [],
                {
                    "aiI`0": ["ai", 2],
                    "auU`0": ["au", 2],
                    "aiI`1": ["ai", 0],
                    "auU`1": ["au", 0],
                },
            ],
            "Purna-Devanagari": [
                [],
                {
                    "aiI`0": ["ai", 2],
                    "auU`0": ["au", 2],
                    "aiI`1": ["ai", 0],
                    "auU`1": ["au", 0],
                    "a1`1": ["a", 2],
                    "A1`1": ["A", 2],
                    "u1`1": ["u", 2],
                    "U1`1": ["U", 2],
                    "g1`0": ["g", 2],
                    "j1`0": ["j", 2],
                    "D1`0": ["D", 2],
                    "b1`0": ["b", 2],
                },
            ],
            "Tamil": [
                ["LR`0", "LRR`0", "R`0", "RR`0"],
                {
                    "க": ["k", 1],
                    "ச": ["ch", 1],
                    "த": ["t", 1],
                    "ட": ["T", 1],
                    "ப": ["p", 1],
                    "ஔ": ["au", 2],
                    "ம்": ["m", 2.1],
                },
            ],
            "Tamil-Extended": [
                ["R`0", "RR`0", "LR`0", "LRR`0"],
                {
                    "ச²": ["chh", 3],
                    "ம்": ["m", 2.1],
                    "ரு": ["ru", 2],
                    "ரூ": ["ru", 2],
                    "R`0": ["R", 2],
                    "RR`0": ["RR", 2],
                    "்ர": ["r", 1.1],
                    "்ரு": ["ru", 2.1],
                    "்ரு²": ["R", 0.1],
                    "்ரூ": ["rU", 2.1],
                    "்ரூ²": ["RR", 0.1],
                    "ல்": ["l", 2.1],
                    "ல்ர": ["lr", 1.1],
                    "ல்ரி": ["lri", 2],
                    "ல்ரீ": ["lrI", 2],
                    "LR`0": ["LR", 2],
                    "LRR`0": ["LRR", 2],
                    "்ல": ["l", 1.1],
                    "்ல்": ["l", 2.1],
                    "்ல்ர": ["lr", 1.1],
                    "்ல்ரி": ["lri", 1.1],
                    "்ல்ரி²": ["LR", 0.1],
                    "்ல்ரீ": ["lrI", 1.1],
                    "்ல்ரீ²": ["LRR", 0.1],
                },
            ],
            "Bengali": [
                [],
                {"b`0": ["b", 1], "t1`0": ["t", 1]},
            ],
            "Siddham": [
                ["LR`0", "LRR`0"],
                {
                    "i1`0": ["i", 2],
                    "i2`0": ["i", 2],
                    "I1`0": ["I", 2],
                    ".u1`0": ["u", 0],
                    ".U1`0": ["U", 0],
                },
            ],
            "Odia": [
                [],
                {
                    "ୱ": ["v", 1],
                    "ଯ": ["y", 1],
                },
            ],
            "Gujarati": [
                [],
                {
                    "aiI`0": ["au", 2],
                    "auU`0": ["ai", 2],
                    "aiI`1": ["au", 0],
                    "auU`1": ["ai", 0],
                },
            ],
            "Assamese": [["rz`0"], {}],
            "Punjabi": [["Dhz`0", "LR`0", "R`0", "LRR`0"], {"ੰ": ["M", 2]}],
            "Urdu": [
                ["ai`0", "au`0", "R`0", "RR`0", "LR`0", "LRR`0"],
                {
                    "a1`0": ["a", 2],
                    "a2`0": ["a", 2],
                    "aa`0": ["a", 2],
                    "i1`0": ["i", 2],
                    "u1`0": ["u", 2],
                    "e1`0": ["e", 2],
                    "U1`0": ["U", 2],
                    "I1`0": ["I", 2],
                    "U2`0": ["U", 2],
                    "I2`0": ["I", 2],
                    "y`0": ["i", 2],
                    "v`0": ["u", 2],
                    "y1`0": ["i", 2],
                    "v1`0": ["u", 2],
                    "h4`0": ["h", 2],
                    "k1`0": ["k", 2],
                    "kh1`0": ["khz", 2],
                    "g1`0": ["gz", 2],
                    "jz1`0": ["jz", 2],
                    "jz2`0": ["jz", 2],
                    "jz3`0": ["jz", 2],
                    "z`0": ["jz", 2],
                    "z1`0": ["jz", 2],
                    "z2`0": ["jz", 2],
                    "z3`0": ["jz", 2],
                    "t1`0": ["t", 2],
                    "t2`0": ["t", 2],
                    "s1`0": ["s", 2],
                    "s2`0": ["s", 2],
                    "h1`0": ["h", 2],
                    "h2`0": ["h", 2],
                    "sh`0": ["sh", 2],
                    "k1h`0": ["kh", 2],
                    "n1`0": ["n", 2],
                    "phz`0": ["phz", 2],
                    "l`0": ["l", 2],
                },
            ],
        }
        for x in ["Hindi", "Nepali", "Marathi", "Konkani", "Sindhi", "Kashmiri"]:
            data[x] = sh.deepcopy(data["Sanskrit"])
        # Addition of .1 after codes are same as alt_codes just the thing is that they are for multiple characters
        for mtr in bhAShAmAtrA[lang]:
            # Alternative numbering for mAtrAs
            # Adding Standard forms of different matra's like A and A1 brahmi are both A
            vl = bhAShAmAtrA[lang][mtr]
            if vl[-1] in "123456":
                if lang not in data:
                    data[lang] = [[], {}]
                data[lang][1][mtr] = [vl[:-1], 0]
                # We are only adding the chars but the real evaluation for their index will be done later
                alt_numb[lang][ak[vl[0]][vl[:-1]][1]] = mtr
        if lang in data:
            dt = data[lang]
            for dl in dt[0]:  # for deletions
                v = dl.split("`")
                if len(v) == 1:
                    del p1[dl]
                elif len(v) == 2:
                    del p1[ak[v[0][0]][v[0]][int(v[1])]]
            for dl in dt[1]:  # for additions
                v = dl.split("`")
                if len(v) == 1:
                    p1[dl] = dt[1][dl]
                elif len(v) == 2:
                    p1[ak[v[0][0]][v[0]][int(v[1])]] = dt[1][dl]
        if lang == "Tamil-Extended":
            for w in "kh,g,gh,chh,jh,th,d,dh,Th,D,Dh,ph,b,bh".split(","):
                vb = ak[w[0]][w][0]
                p1[vb[:-1] + {"²": "₂", "³": "₃", "⁴": "₄"}[vb[-1]]] = [w, 3]
        for x in "²³⁴₂₃₄":
            if x in p1:
                del p1[x]
        if ".x" in ak["."]:
            if ak["."][".x"][0] in p1:
                p1[ak["."][".x"][0]][0] = ""
        p1["↓"] = ["#an", 2]
        p1["↑"] = ["#s", 2]
        p1["↑↑"] = ["#ss", 2]
        p1["↑↑↑"] = ["#sss", 2]
        if "" in p1:
            del p1[""]
        if "AUM1" in ak["A"]:
            if ak["A"]["AUM1"][0] == "ॐ":
                p1["ॐ"] = ["AUM", 2]
        # adding nukta flavour
        if lang in ("Romanized", "Normal", "Urdu"):
            continue
        if ".z" in ak["."] and lang not in ("Granth", "Siddham"):
            nukta = ak["."][".z"][0]
            p1[ak["p"]["ph"][0] + nukta] = ["phz", 1]
            p1[ak["p"]["phz"][0]] = ["phz", 1]
    p1 = sh.deepcopy(p)
    # diifferent as we had to deal with p and p1 seperately
    # to construct p1 values of p are needed here
    for lang in p:
        if lang in ("Romanized", "Normal"):
            continue
        if ".z" not in a[lang]["."]:
            continue
        # Adding alt_codes for nukta based alternatives
        for varna in p[lang]:
            v = p[lang][varna]
            if v[1] == 1 and v[0][-1] == "z" and len(varna) == 1:
                tr = v[0][:-1]
                fg = a[lang][tr[0]][tr][0] + a[lang]["."][".z"][0]
                p1[lang][fg] = [v[0], v[1]]
                # We are only adding the chars but the real evaluation for their index will be done later
                alt_numb[lang][varna] = fg
    kram = {}
    # adding varna kram
    for cl in range(1, sht.max_row + 1):
        l = sht.cell(1, cl)._value
        if cl == 1:
            l = "Purna-Devanagari"
        if l == None:
            break
        kram[l] = []
        for rw in range(2, sht.max_row + 1):
            vl = sht.cell(rw, cl)._value
            if vl == None:
                break
            if vl not in (0, 1) and cl != 1:
                if vl[0] == "`":
                    vl = vl[1:]
                gh = ""
                for t in vl.split("+"):
                    gh += a[l][t[0]][t][0]
                vl = gh
            kram[l].append(vl)
    # adding continuations
    for l in kram:
        # ^ scanning for which are there in kram but not in antar
        for x in kram[l]:  # Use Not identified yet
            if l not in ["Tamil-Extended"]:
                continue
            if x not in p1[l] and x not in (0, 1):
                print("ok1", l, x)
                k1 = kram["Purna-Devanagari"][kram[l].index(x)]
                k = p1["Purna-Devanagari"][k1][1]
                sh.clip_copy(f'"{x}": ["", {k}],')
                input(f'{l} => "{x}": {list(k1)} ["", {k}],')
    p = sh.deepcopy(p1)
    for l in p:  # this is actually adding continuations
        for v in p[l]:
            if len(v) == 1:
                continue
            if v[:-1] in p1[l]:
                el = p1[l][v[:-1]]
                if len(el) == 2:
                    el.append(v[-1])
                else:
                    el[2] += v[-1]
            else:
                print(f"{l}   {v}    >> {p1[l][v]}, {len(v)}")
    # sorting
    p2 = {}
    for l in p1:
        p2[l] = {}
        for v in sorted(p1[l].keys()):
            if v == p1[l][v][0]:
                continue
            p2[l][v] = p1[l][v]
    p = sh.deepcopy(p2)
    for l in alt_numb:  # giving the alternative numbers
        for x in alt_numb[l]:
            if x in kram[l]:
                vl = p[l][alt_numb[l][x]]
                if len(vl) == 2:
                    vl.append("")
                    vl.append(kram[l].index(x))
                else:
                    vl.append(kram[l].index(x))
    a2 = sh.deepcopy(a)
    for x in a:
        a2[x]["antar"] = p[x]
    for x in kram:
        a2[x]["kram"] = kram[x]
    a = a2
    return a
